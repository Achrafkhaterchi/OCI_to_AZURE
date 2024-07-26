import os
import pandas as pd
from openpyxl import load_workbook
from New_Subnet import new_subnet, subnetting
from jinja2 import Environment, FileSystemLoader, select_autoescape


parent_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(parent_dir, 'OCI.xlsx')
template_file_path = os.path.join(parent_dir, '..', 'Network', 'VNets', 'Template', 'VNets.j2')
output_file_path = os.path.join(parent_dir, '..', 'Network', 'VNets', 'variables.tfvars')

wb = load_workbook(excel_file_path, data_only=True)

#--------------------------------------------------------------------------------------------------------------------------

def extract_resource_group(wb):
    vcn_sheet = wb["RG"]
    resource_group = vcn_sheet.cell(row=2, column=1).value
    return resource_group

resource_group = extract_resource_group(wb)

#--------------------------------------------------------------------------------------------------------------------------

def extract_region(wb):
    compartment_sheet = wb["Region"]
    region = compartment_sheet.cell(row=2, column=1).value
    return region

region = extract_region(wb)

#--------------------------------------------------------------------------------------------------------------------------

def extract_vcns(wb):
    vcn_sheet = wb["VCN"]
    vcns = {}
    gateway_vnet_name = None
    for row in vcn_sheet.iter_rows(min_row=2, values_only=True):
        compartment_name, vcn_name, dns_label, cidr_block = row
        if vcn_name not in vcns:
            vcns[vcn_name] = {"address_space": [cidr_block]}
        if dns_label == "hub":
            gateway_vnet_name = vcn_name
    return vcns, gateway_vnet_name

vcns, gateway_vnet_name = extract_vcns(wb)

#--------------------------------------------------------------------------------------------------------------------------

def extract_all_subnets(wb):
    subnet_sheet = wb["Subnets"]
    all_subnets = []
    for row in subnet_sheet.iter_rows(min_row=2, values_only=True):
        vcn_name, subnet_name, subnet_cidr, _ = row
        all_subnets.append({
            "vcn_name": vcn_name,
            "subnet_name": subnet_name,
            "subnet_cidr": subnet_cidr
        })
    return all_subnets

all_subnets = extract_all_subnets(wb)

#--------------------------------------------------------------------------------------------------------------------------

def extract_hub_subnets(wb, vcns, gateway_vnet_name):
    subnet_sheet = wb["Subnets"]
    subnets = {}
    for row in subnet_sheet.iter_rows(min_row=2, values_only=True):
        vcn_name, subnet_name, subnet_cidr, _ = row
        if vcn_name in vcns and vcn_name == gateway_vnet_name:
            if vcn_name not in subnets:
                subnets[vcn_name] = []
            subnets[vcn_name].append(subnet_cidr)
    return subnets

subnets = extract_hub_subnets(wb, vcns, gateway_vnet_name)

#--------------------------------------------------------------------------------------------------------------------------

gateway_subnet = {}
last_subnet = False

if gateway_vnet_name and gateway_vnet_name in vcns and gateway_vnet_name in subnets:
    vnet_prefix = vcns[gateway_vnet_name]["address_space"][0]
    subnets_prefixes = subnets[gateway_vnet_name]
    last_subnet = new_subnet(vnet_prefix, subnets_prefixes)
    if last_subnet:
        gateway_subnet = {
            "vnet_name": gateway_vnet_name,
            "address_prefixes": str(last_subnet)
        }


if not last_subnet:
    vnet_address_space = vcns[gateway_vnet_name]["address_space"][0]
    num_subnets = len(subnets[gateway_vnet_name]) + 2
    subnetting_result = subnetting(vnet_address_space, num_subnets)
    gateway_subnet = {
        "vnet_name": gateway_vnet_name,
        "address_prefixes": str(subnetting_result[0])  
    }
    for i, subnet in enumerate(all_subnets):
        if subnet["vcn_name"] == gateway_vnet_name:
            subnet["subnet_cidr"] = subnetting_result[i + 1]

#--------------------------------------------------------------------------------------------------------------------------
            
template_data = {
    "resource_group": resource_group,
    "region": region,
    "vnets": vcns,
    "GatewaySubnet": gateway_subnet,
    "subnets": all_subnets
}

env = Environment(
    loader=FileSystemLoader(os.path.dirname(template_file_path)),
    autoescape=select_autoescape(['j2'])
)
template = env.get_template(os.path.basename(template_file_path))

rendered_content = template.render(template_data)

with open(output_file_path, 'w') as f:
    f.write(rendered_content)

print("Fichier variables.tfvars généré avec succès.")
