from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import pandas as pd
import json
import os


parent_dir = os.path.dirname(os.path.abspath(__file__))
tfstate_file_path = os.path.join(parent_dir, 'network-tfstate_terraform.tfstate')
tfstate_iam_path = os.path.join(parent_dir, 'terraform.tfstate')
excel_file_path = os.path.join(parent_dir, 'OCI.xlsx')


with open(tfstate_file_path, "r") as file:
    tfstate_data = json.load(file)
    
with open(tfstate_iam_path, "r") as file:
    tfstate_iam_data = json.load(file)
    
wb = load_workbook(excel_file_path)

compartments_sheet = wb["Compartments"]
groups_sheet = wb["Groups"]
policies_sheet = wb["Policies"]


#------------------------------------------------------------Name_Id-------------------------------------------------------------
compartment_dict = {
    compartment["id"]: compartment["name"]
    for compartment in tfstate_iam_data["outputs"]["compartments"]["value"].values()
}

vcn_dict = {
    instance["attributes"]["id"]: instance["attributes"]["display_name"]
    for resource in tfstate_data["resources"]
    if "type" in resource and resource["type"] == "oci_core_vcn"
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", [])
    if "attributes" in instance and "id" in instance["attributes"] and "display_name" in instance["attributes"]
}

Route_Table_dict = {
    instance["attributes"]["id"]: instance["attributes"]["display_name"]
    for resource in tfstate_data["resources"]
    if "type" in resource and resource["type"] == "oci_core_route_table"
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", [])
    if "attributes" in instance and "id" in instance["attributes"] and "display_name" in instance["attributes"]
}

Subnet_dict = {
    instance["attributes"]["id"]: instance["attributes"]["display_name"]
    for resource in tfstate_data["resources"]
    if "type" in resource and resource["type"] == "oci_core_subnet"
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", [])
    if "attributes" in instance and "id" in instance["attributes"] and "display_name" in instance["attributes"]
}

nsg_dict = {
    instance["attributes"]["id"]: instance["attributes"]["display_name"]
    for resource in tfstate_data["resources"]
    if "type" in resource and resource["type"] == "oci_core_network_security_group"
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", [])
    if "attributes" in instance and "id" in instance["attributes"] and "display_name" in instance["attributes"]
}

security_list_dict = {
    instance["attributes"]["id"]: instance["attributes"]["display_name"]
    for resource in tfstate_data["resources"]
    if "type" in resource and resource["type"] == "oci_core_security_list"
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", [])
    if "attributes" in instance and "id" in instance["attributes"] and "display_name" in instance["attributes"]
} 

protocol_mapping = {
    '6': 'Tcp',
    '17': 'Udp',
    '1': 'Icmp'
}



region_map = {
    "eu-paris-1": "eastus"
}

#------------------------------------------------------------Region---------------------------------------------------------------

region_ = [
    instance["attributes"]["reporting_region"]
    for resource in tfstate_iam_data["resources"]
    if "type" in resource and resource["type"] == "oci_cloud_guard_cloud_guard_configuration"
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", [])
    if "attributes" in instance and "reporting_region" in instance["attributes"]
]

region = ', '.join(region_map.get(r, r) for r in region_)

sheet_name = "Region"
ws = wb[sheet_name]

row = ws.max_row + 1
ws[f"A{row}"] = region

wb.save(excel_file_path)
wb.close()

#--------------------------------------------------------------VCNs---------------------------------------------------------------

grouped_vcn_data_list = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_vcn" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            attributes = instance.get("attributes", {})
            compartment_id = attributes.get("compartment_id", "Unknown ID")
            compartment_name = compartment_dict.get(compartment_id, "Unknown ID")
            vcn_name = attributes.get("display_name", "Unknown VCN")
            dns_label = attributes.get("dns_label", "")
            cidr_block = attributes.get("cidr_block", "")

            for grouped_data in grouped_vcn_data_list:
                if grouped_data["Compartment_ID"] == compartment_id:
                    grouped_data["VCN_Name"].append(vcn_name)
                    grouped_data["DNS_Label"].append(dns_label)
                    grouped_data["CIDR_Block"].append(cidr_block)
                    break
            else:
                grouped_vcn_data_list.append({
                    "Compartment_ID": compartment_id,
                    "Compartment_Name": compartment_name,
                    "VCN_Name": [vcn_name],
                    "DNS_Label": [dns_label],
                    "CIDR_Block": [cidr_block]
                })

for data in grouped_vcn_data_list:
    data.pop("Compartment_ID")

df_vcn = pd.DataFrame(grouped_vcn_data_list)
sheet_name = "VCN"
ws = wb[sheet_name]

sheet_rg_name = "RG"
ws_rg = wb[sheet_rg_name]


for index, row in df_vcn.iterrows():
    compartment_name = row["Compartment_Name"]
    vcn_names = row["VCN_Name"]
    dns_labels = row["DNS_Label"]
    cidr_blocks = row["CIDR_Block"]

    for vcn_name, dns_label, cidr_block in zip(vcn_names, dns_labels, cidr_blocks):
        ws.append([compartment_name, vcn_name, dns_label, cidr_block])

for index, row in df_vcn.iterrows():
    compartment_name = row["Compartment_Name"]
    ws_rg.append([compartment_name])

wb.save(excel_file_path)
wb.close()


#--------------------------------------------------------------Subnets------------------------------------------------------------

grouped_subnet_data_list = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_subnet" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            attributes = instance.get("attributes", {})
            vcn_id = attributes.get("vcn_id", "Unknown ID")
            vcn_name = vcn_dict.get(vcn_id, "Unknown ID")
            subnet_name = attributes.get("display_name", "Unknown Subnet")
            subnet_cidr = attributes.get("cidr_block", "")
            subnet_dns_label = attributes.get("dns_label", "")

            grouped_subnet_data_list.append({
                "VCN_Name": vcn_name,
                "Subnet_Name": subnet_name,
                "Subnet_CIDR": subnet_cidr,
                "Subnet_DNS_Label": subnet_dns_label
            })

df2 = pd.DataFrame(grouped_subnet_data_list)

df2.sort_values(by=["VCN_Name"], inplace=True)

sheet_name = "Subnets"
ws = wb[sheet_name]

for index, row in df2.iterrows():
    vcn_name = row["VCN_Name"]
    subnet_name = row["Subnet_Name"]
    subnet_cidr = row["Subnet_CIDR"]
    subnet_dns_label = row["Subnet_DNS_Label"]

    ws.append([vcn_name, subnet_name, subnet_cidr, subnet_dns_label])

wb.save(excel_file_path)
wb.close()


#--------------------------------------------------------------Route Tables---------------------------------------------------------

grouped_route_data_list = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_route_table" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance and "display_name" in instance["attributes"]:
                route_table_name = instance["attributes"]["display_name"]
                vcn_id = instance["attributes"]["vcn_id"]
                vcn_name = vcn_dict.get(vcn_id, "Unknown ID")

                route_rule_data = {
                    "VCN_Name": vcn_name,
                    "Route_Table_Name": route_table_name,
                    "Route_Rule_Description": [],
                    "Route_Rule_Destination": [],
                    "Route_Rule_Destination_Type": []
                }

                if "route_rules" in instance["attributes"]:
                    for route_rule in instance["attributes"]["route_rules"]:
                        if "description" in route_rule:
                            route_rule_data["Route_Rule_Description"].append(route_rule["description"])
                        if "destination" in route_rule:
                            route_rule_data["Route_Rule_Destination"].append(route_rule["destination"])
                        if "destination_type" in route_rule:
                            route_rule_data["Route_Rule_Destination_Type"].append(route_rule["destination_type"])

                grouped_route_data_list.append(route_rule_data)

df_route_tables = pd.DataFrame(grouped_route_data_list)

sheet_name = "Route_Table"
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

for index, row in df_route_tables.iterrows():
    vcn_name = row["VCN_Name"]
    route_table_name = row["Route_Table_Name"]
    route_rule_descriptions = row["Route_Rule_Description"]
    route_rule_destinations = row["Route_Rule_Destination"]
    route_rule_destination_types = row["Route_Rule_Destination_Type"]

    for description, destination, destination_type in zip(route_rule_descriptions, route_rule_destinations, route_rule_destination_types):
        ws.append([vcn_name, route_table_name, description, destination, destination_type])

wb.save(excel_file_path)
wb.close()

#--------------------------------------------------------------Route Tables & Subnets---------------------------------------------------------

Route_Table_ID = []
Subnet_Name_ID = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_route_table_attachment" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance:
                if "route_table_id" in instance["attributes"]:
                    Route_Table_ID.append(instance["attributes"]["route_table_id"])
                if "subnet_id" in instance["attributes"]:
                    Subnet_Name_ID.append(instance["attributes"]["subnet_id"])

Route_Table_Name_map = [Route_Table_dict.get(rt_id, "Unknown ID") for rt_id in Route_Table_ID]
Subnet_Name_map = [Subnet_dict.get(sb_id, "Unknown ID") for sb_id in Subnet_Name_ID]

rt_attachment_data = {
    "Route_Table_Name": Route_Table_Name_map,
    "Subnet_Name": Subnet_Name_map
}
df_rt_attachment = pd.DataFrame(rt_attachment_data)

sheet_name = "RT_Attachment"
ws = wb[sheet_name]

rows = dataframe_to_rows(df_rt_attachment, index=False, header=True) 

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)
wb.close()


#--------------------------------------------------------------Nat Gateway-----------------------------------------------------------------

Nat_GW = []
Block_Trafic = []
Nat_IP = []
VCN_IDs = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_nat_gateway" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance:
                if "display_name" in instance["attributes"]:
                    Nat_GW.append(instance["attributes"]["display_name"])
                if "block_traffic" in instance["attributes"]:
                    Block_Trafic.append(str(instance["attributes"]["block_traffic"]).lower())
                if "nat_ip" in instance["attributes"]:
                    Nat_IP.append(instance["attributes"]["nat_ip"])
                if "vcn_id" in instance["attributes"]:
                    VCN_IDs.append(instance["attributes"]["vcn_id"])

VCN_Names = [vcn_dict.get(vcn_id, "Unknown ID") for vcn_id in VCN_IDs]

nat_gateway_data = {
    "VCN_Name": VCN_Names,
    "NAT_GW_Name": Nat_GW,
    "Block_Traffic": Block_Trafic,
    "NAT_IP": Nat_IP
}
df_nat_gateway = pd.DataFrame(nat_gateway_data)

sheet_name = "NAT"
ws = wb[sheet_name]

rows = dataframe_to_rows(df_nat_gateway, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)
wb.close()



#--------------------------------------------------------------Internet Gateway-----------------------------------------------------------------

Internet_GW = []
VCN_IDs = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_internet_gateway" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance:
                if "display_name" in instance["attributes"]:
                    Internet_GW.append(instance["attributes"]["display_name"])
                if "vcn_id" in instance["attributes"]:
                    VCN_IDs.append(instance["attributes"]["vcn_id"])

VCN_Names = [vcn_dict.get(vcn_id, "Unknown ID") for vcn_id in VCN_IDs]

internet_gw_data = {
    "Internet_GW_Name": Internet_GW,
    "VCN_Name": VCN_Names
}
df_internet_gw = pd.DataFrame(internet_gw_data)

sheet_name = "IGW"
ws = wb[sheet_name]

rows = dataframe_to_rows(df_internet_gw, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)
wb.close()

#---------------------------------------------------------------Service Gateway-------------------------------------------------------------------

Service_GW = []
VCN_IDs_gw = []
Block_Trafic = []
Service_Name = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_service_gateway" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance:
                if "display_name" in instance["attributes"]:
                    Service_GW.append(instance["attributes"]["display_name"])
                if "vcn_id" in instance["attributes"]:
                    VCN_IDs_gw.append(instance["attributes"]["vcn_id"])
                if "block_traffic" in instance["attributes"]:
                    Block_Trafic.append(str(instance["attributes"]["block_traffic"]).lower())
                if "services" in instance["attributes"]:
                    for service in instance["attributes"]["services"]:
                        if "service_name" in service:
                            Service_Name.append(service["service_name"])
                


VCN_Names_gw = [vcn_dict.get(vcn_id, "Unknown ID") for vcn_id in VCN_IDs_gw]

service_gw_data = {
    "VCN_Name": VCN_Names_gw,
    "Service_Name": Service_Name,
    "Service_GW_Name": Service_GW,
    "Block Trafic": Block_Trafic
}
df_service_gw = pd.DataFrame(service_gw_data)

sheet_name = "SGW"
ws = wb[sheet_name]

rows = dataframe_to_rows(df_service_gw, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)
wb.close()



#---------------------------------------------------------------Network Security Groups-------------------------------------------------------------------
grouped_nsg_data_list = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_network_security_group" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance and "display_name" in instance["attributes"]:
                vcn_id = instance["attributes"]["vcn_id"]
                vcn_name = vcn_dict.get(vcn_id, "Unknown ID")
                nsg_name = instance["attributes"]["display_name"]
                
                existing_entry = next((item for item in grouped_nsg_data_list if item["VCN_Name"] == vcn_name), None)
                
                if existing_entry:
                    existing_entry["NSG_Name"].append(nsg_name)
                else:
                    nsg_data = {
                        "VCN_Name": vcn_name,
                        "NSG_Name": [nsg_name]
                    }
                    grouped_nsg_data_list.append(nsg_data)

df_service_nsg = pd.DataFrame(grouped_nsg_data_list)

sheet_name = "NSG"

ws = wb[sheet_name]

rows = dataframe_to_rows(df_service_nsg, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, start=1):
        if isinstance(value, list):
            value = '\n'.join(value)
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)

wb.close()


#--------------------------------------------------------------------------NSG Rules-------------------------------------------------------------------

grouped_nsg_data_list = []
ocid_prefix = "ocid1.networksecuritygroup"
for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_network_security_group_security_rule" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance:
                nsg_id = instance["attributes"]["network_security_group_id"]
                nsg_name = nsg_dict.get(nsg_id, "Unknown NSG ID")
                rule_description = instance["attributes"]["description"]
                destination = instance["attributes"]["destination"]

                if destination and destination.startswith(ocid_prefix):
                    destination = nsg_dict.get(destination, "Unknown ID")

                    
                destination_type = instance["attributes"]["destination_type"]
                direction = instance["attributes"]["direction"]

                protocol_num = instance["attributes"]["protocol"]
                protocol = protocol_mapping.get(protocol_num, "Unknown")
                
                source = instance["attributes"]["source"]

                if source and source.startswith(ocid_prefix):
                    source = nsg_dict.get(source, "Unknown ID")
                source_type = instance["attributes"]["source_type"]
                stateless = instance["attributes"]["stateless"]
                
                tcp_options = instance["attributes"].get("tcp_options", [{}])
                destination_port_range_str = "N/A"
                source_port_range_str = "N/A"
                
                if tcp_options and "destination_port_range" in tcp_options[0]:
                    destination_port_ranges = tcp_options[0]["destination_port_range"]
                    if destination_port_ranges:
                        destination_port_range_str = ', '.join(f"{port['min']}-{port['max']}" for port in destination_port_ranges)
                
                if tcp_options and "source_port_range" in tcp_options[0]:
                    source_port_ranges = tcp_options[0]["source_port_range"]
                    if source_port_ranges:
                        source_port_range_str = ', '.join(f"{port['min']}-{port['max']}" for port in source_port_ranges)
                
                existing_entry = next((item for item in grouped_nsg_data_list if item["NSG_Name"] == nsg_name), None)
                
                if existing_entry:
                    existing_entry["Description"].append(rule_description)
                    existing_entry["Destination"].append(destination)
                    existing_entry["Destination_Type"].append(destination_type)
                    existing_entry["Direction"].append(direction)
                    existing_entry["Protocol"].append(protocol)
                    existing_entry["Source"].append(source)
                    existing_entry["Source_Type"].append(source_type)
                    existing_entry["Stateless"].append(stateless)
                    existing_entry["Destination_Port_Range"].append(destination_port_range_str)
                    existing_entry["Source_Port_Range"].append(source_port_range_str)
                else:
                    rule_data = {
                        "NSG_Name": nsg_name,
                        "Description": [rule_description],
                        "Destination": [destination],
                        "Destination_Type": [destination_type],
                        "Direction": [direction],
                        "Protocol": [protocol],
                        "Source": [source],
                        "Source_Type": [source_type],
                        "Stateless": [stateless],
                        "Destination_Port_Range": [destination_port_range_str],
                        "Source_Port_Range": [source_port_range_str]
                    }
                    grouped_nsg_data_list.append(rule_data)

df_nsg = pd.DataFrame(grouped_nsg_data_list)

sheet_name = "NSG_Rules"
ws = wb[sheet_name]

rows = dataframe_to_rows(df_nsg, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, start=1):
        if isinstance(value, list):
            value = '\n'.join(map(str, value))
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)
wb.close()

#---------------------------------------------------------------Security List Rules-----------------------------------------------------------------

security_list_rules_data_list = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_security_list" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance:
                attributes = instance["attributes"]
                display_name = attributes.get("display_name")
                
                for rule in attributes.get("ingress_security_rules", []):
                    protocol = rule.get("protocol")
                    protocol_name = protocol_mapping.get(protocol, protocol)
                    
                    security_list_rules_data_list.append({
                        "display_name": display_name,
                        "direction": "ingress",
                        "description": rule.get("description"),
                        "protocol": protocol_name,
                        "icmp_options": rule.get("icmp_options"),
                        "source": rule.get("source"),
                        "source_type": rule.get("source_type"),
                        "destination": None,
                        "destination_type": None,
                        "tcp_options": rule.get("tcp_options"),
                        "udp_options": rule.get("udp_options")
                    })
                
                for rule in attributes.get("egress_security_rules", []):
                    protocol = rule.get("protocol")
                    protocol_name = protocol_mapping.get(protocol, protocol)
                    
                    security_list_rules_data_list.append({
                        "display_name": display_name,
                        "direction": "egress",
                        "description": rule.get("description"),
                        "protocol": protocol_name,
                        "icmp_options": rule.get("icmp_options"),
                        "source": None,
                        "source_type": None,
                        "destination": rule.get("destination"),
                        "destination_type": rule.get("destination_type"),
                        "tcp_options": rule.get("tcp_options"),
                        "udp_options": rule.get("udp_options")
                    })

df_security_list_rules = pd.DataFrame(security_list_rules_data_list)

sheet_name = "security_list_rules"
ws = wb[sheet_name]
rows = dataframe_to_rows(df_security_list_rules, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, start=1):
        if isinstance(value, list):
            value = '\n'.join(map(str, value))
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)

wb.close()

#---------------------------------------------------------------Security List Associations------------------------------------------------------------

subnet_security_list_association = []

for resource in tfstate_data.get("resources", []):
    if resource.get("type") == "oci_core_subnet" and resource.get("mode") == "managed":
        for instance in resource.get("instances", []):
            if "attributes" in instance:
                attributes = instance["attributes"]
                display_name = attributes.get("display_name")
                security_list_ids = attributes.get("security_list_ids", [])
                
                for sec_list_id in security_list_ids:
                    security_list_name = security_list_dict.get(sec_list_id)
                    if security_list_name:
                        subnet_security_list_association.append({
                            "Subnet_Name": display_name,
                            "Security_List_Name": security_list_name
                        })

df_subnet_security_list_association = pd.DataFrame(subnet_security_list_association)


sheet_name = "Security_List_Associations"
ws = wb[sheet_name]

rows = dataframe_to_rows(df_subnet_security_list_association, index=False, header=True)

for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, start=1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(excel_file_path)

wb.close()









