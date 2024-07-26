import os
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader, select_autoescape

parent_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(parent_dir, 'OCI.xlsx')
template_file_path = os.path.join(parent_dir, '..', 'Network', 'Security', 'Template', 'Security.j2')
output_file_path = os.path.join(parent_dir, '..', 'Network', 'Security', 'variables.tfvars')

wb = load_workbook(excel_file_path, data_only=True)

#----------------------------------------------------------------------------------------------------------------------------

def extract_resource_group(wb):
    rg_sheet = wb["RG"]
    resource_group = rg_sheet.cell(row=2, column=1).value
    return resource_group

resource_group = extract_resource_group(wb)

#----------------------------------------------------------------------------------------------------------------------------

def extract_region(wb):
    compartment_sheet = wb["Region"]
    region = compartment_sheet.cell(row=2, column=1).value
    return region

region = extract_region(wb)

#----------------------------------------------------------------------------------------------------------------------------

def extract_security_list_rules(wb):
    security_list_rules_sheet = wb["security_list_rules"]
    rules_by_nsg = {}
    priority = 100
    display_name_counter = {}

    for row in security_list_rules_sheet.iter_rows(min_row=2, values_only=True):
        display_name, direction, description, protocol, icmp_options, source, source_type, destination, destination_type, tcp_options, udp_options = row

        direction = "Inbound" if direction.lower() == "ingress" else "Outbound"
        access = "Allow"

        if display_name not in display_name_counter:
            display_name_counter[display_name] = 0
        display_name_counter[display_name] += 1
        rule_display_name = f"rule{display_name_counter[display_name]}"

        rule = {
            "name": rule_display_name,
            "priority": priority,
            "direction": direction,
            "access": access,
            "protocol": protocol,
            "source_port_range": "*",
            "destination_port_range": "*",
            "source_address_prefix": source if source_type == "CIDR_BLOCK" and direction == "Inbound" else "*",
            "destination_address_prefix": destination if destination_type == "CIDR_BLOCK" and direction == "Outbound" else "*"
        }

        if protocol == "TCP" and tcp_options:
            tcp_options = eval(tcp_options)
            if direction == "Inbound":
                if tcp_options["max"] == tcp_options["min"]:
                    rule["source_port_range"] = str(tcp_options["max"])
                else:
                    rule["source_port_range"] = f"{tcp_options['min']}-{tcp_options['max']}"
            else:
                if tcp_options["max"] == tcp_options["min"]:
                    rule["destination_port_range"] = str(tcp_options["max"])
                else:
                    rule["destination_port_range"] = f"{tcp_options['min']}-{tcp_options['max']}"
        elif protocol == "UDP" and udp_options:
            udp_options = eval(udp_options)
            if direction == "Inbound":
                if udp_options["max"] == udp_options["min"]:
                    rule["source_port_range"] = str(udp_options["max"])
                else:
                    rule["source_port_range"] = f"{udp_options['min']}-{udp_options['max']}"
            else:
                if udp_options["max"] == udp_options["min"]:
                    rule["destination_port_range"] = str(udp_options["max"])
                else:
                    rule["destination_port_range"] = f"{udp_options['min']}-{udp_options['max']}"
        elif protocol == "ICMP":
            rule["source_port_range"] = "*"
            rule["destination_port_range"] = "*"

        if display_name not in rules_by_nsg:
            rules_by_nsg[display_name] = []
        rules_by_nsg[display_name].append(rule)
        priority += 10

    return rules_by_nsg

security_list_rules = extract_security_list_rules(wb)

#----------------------------------------------------------------------------------------------------------------------------

def extract_security_list_associations(wb):
    security_list_associations_sheet = wb["Security_List_Associations"]
    associations = []

    for row in security_list_associations_sheet.iter_rows(min_row=2, values_only=True):
        subnet_name, security_list_name = row
        associations.append({
            "subnet_name": subnet_name,
            "security_list_name": security_list_name
        })

    return associations

security_list_associations = extract_security_list_associations(wb)

#----------------------------------------------------------------------------------------------------------------------------

template_data = {
    "resource_group": resource_group,
    "region": region,
    "network_security_groups": security_list_rules,
    "subnet_nsg_associations": security_list_associations
}

env = Environment(
    loader=FileSystemLoader(os.path.dirname(template_file_path)),
    autoescape=select_autoescape(['j2'])
)
template = env.get_template(os.path.basename(template_file_path))

rendered_content = template.render(template_data)

with open(output_file_path, 'w') as f:
    f.write(rendered_content)

print("Fichier variables.tfvars généré avec succès.")
