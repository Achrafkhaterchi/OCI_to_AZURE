from openpyxl import load_workbook
#import oci
import json
import os


parent_dir = os.path.dirname(os.path.abspath(__file__))
tfstate_file_path = os.path.join(parent_dir, 'terraform.tfstate')
excel_file_path = os.path.join(parent_dir, 'OCI.xlsx')

#config = oci.config.from_file()

with open(tfstate_file_path, "r") as file:
    tfstate_data = json.load(file)

wb = load_workbook(excel_file_path)

compartments_sheet = wb["Compartments"]
groups_sheet = wb["Groups"]
policies_sheet = wb["Policies"]

U_id = "fb9d71eb-abe1-41ee-a7f9-23530aaaee3d"


region_ = [
    instance["attributes"]["reporting_region"] 
    for resource in tfstate_data["resources"] 
    if "type" in resource and resource["type"] == "oci_cloud_guard_cloud_guard_configuration" 
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", []) 
    if "attributes" in instance and "reporting_region" in instance["attributes"]
]

region_map = {
    "eu-paris-1": "eastus"
}

region = ', '.join(region_map.get(r, r) for r in region_)


compartments_names = [value["name"] for value in tfstate_data["outputs"]["compartments"]["value"].values()]
compartments_descriptions= [value["description"] for value in tfstate_data["outputs"]["lz_compartments"]["value"]["compartments"].values()]
#print(compartments_names)
#print(compartments_descriptions)

for name, description in zip(compartments_names, compartments_descriptions):
    compartments_sheet.append([region, name, description])


groups_name = [
    instance["attributes"]["name"] 
    for resource in tfstate_data["resources"] 
    if "type" in resource and resource["type"] == "oci_identity_group" 
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", []) 
    if "attributes" in instance and "name" in instance["attributes"]
]

groups_description = [
    instance["attributes"]["description"] 
    for resource in tfstate_data["resources"] 
    if "type" in resource and resource["type"] == "oci_identity_group" 
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", []) 
    if "attributes" in instance and "description" in instance["attributes"]
]



for name, description in zip(groups_name, groups_description):
    groups_sheet.append([region, name, description, U_id])


policy_name = [
    instance["attributes"]["name"] 
    for resource in tfstate_data["resources"] 
    if "type" in resource and resource["type"] == "oci_identity_policy" 
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", []) 
    if "attributes" in instance and "name" in instance["attributes"]
]

policy_description = [
    instance["attributes"]["description"] 
    for resource in tfstate_data["resources"] 
    if "type" in resource and resource["type"] == "oci_identity_policy" 
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", []) 
    if "attributes" in instance and "description" in instance["attributes"]
]

policy_statements = [
    instance["attributes"]["statements"] 
    for resource in tfstate_data["resources"] 
    if "type" in resource and resource["type"] == "oci_identity_policy" 
    and "mode" in resource and resource["mode"] == "managed"
    for instance in resource.get("instances", []) 
    if "attributes" in instance and "statements" in instance["attributes"]
]


row_idx = policies_sheet.max_row + 1 

for name, description, statements in zip(policy_name, policy_description, policy_statements):
    start_row = row_idx
    policies_sheet.append([region, name, description, statements[0]])
    row_idx += 1
    
    for statement in statements[1:]:
        policies_sheet.append(["", "", "", statement])
        row_idx += 1
    
    if len(statements) > 1:
        policies_sheet.merge_cells(start_row=start_row, start_column=1, end_row=row_idx-1, end_column=1)
        policies_sheet.merge_cells(start_row=start_row, start_column=2, end_row=row_idx-1, end_column=2)
        policies_sheet.merge_cells(start_row=start_row, start_column=3, end_row=row_idx-1, end_column=3)




wb.save(excel_file_path)
