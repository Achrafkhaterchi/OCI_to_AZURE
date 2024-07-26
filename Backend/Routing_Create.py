import os
import pandas as pd
from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader, select_autoescape


parent_dir = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(parent_dir, 'OCI.xlsx')
template_file_path = os.path.join(parent_dir, '..', 'Network', 'Routing', 'Template', 'Routing.j2')
output_file_path = os.path.join(parent_dir, '..', 'Network', 'Routing', 'variables.tfvars')

wb = load_workbook(excel_file_path, data_only=True)

#--------------------------------------------------------------------------------------------------------------------------

def extract_resource_group(wb):
    vcn_sheet = wb["RG"]
    resource_group = vcn_sheet.cell(row=2, column=1).value
    return resource_group

resource_group = extract_resource_group(wb)

#--------------------------------------------------------------------------------------------------------------------------

def extract_region(wb):
    compartment_sheet = wb["Region"]
    region = compartment_sheet.cell(row=2, column=1).value
    return region

region = extract_region(wb)

#--------------------------------------------------------------------------------------------------------------------------

def peering(wb):
    vcn_sheet = wb["VCN"]
    hub_vcn_name = None
    spoke_vcns = []

    for row in vcn_sheet.iter_rows(min_row=2, values_only=True):
        _, vcn_name, dns_label, _ = row  
        
        if dns_label == "hub" and hub_vcn_name is None:
            hub_vcn_name = vcn_name
        elif dns_label != "hub":
            spoke_vcns.append(vcn_name)

    return hub_vcn_name, spoke_vcns

hub_vcn_name, spoke_vcns = peering(wb)

#--------------------------------------------------------------------------------------------------------------------------

def extract_route_tables(excel_file_path, sheet_name="Route_Table"):
    df_routes = pd.read_excel(excel_file_path, sheet_name=sheet_name)
    df_vcns = pd.read_excel(excel_file_path, sheet_name="VCN")
    
    route_table_dict = {}
    
    for index, row in df_routes.iterrows():
        table_name = row['Route_Table_Name']
        vcn_name = row['VCN_Name']
        rule_dest = row['Route_Rule_Destination']
        rule_dest_type = row['Route_Rule_Destination_Type']
        
        if rule_dest_type != "CIDR_BLOCK":
            continue
        
        if rule_dest == "0.0.0.0/0":
            next_hop_type = "Internet"
            rule_desc = f"{vcn_name}_To_Internet"
        else:
            next_hop_type = "VirtualNetworkGateway"
            rule_desc = f"{vcn_name}_To_{find_destination_vcn(rule_dest, df_vcns)}"
        
        if table_name not in route_table_dict:
            route_table_dict[table_name] = []
        
        route_table_dict[table_name].append({
            'name': rule_desc,
            'address_prefix': rule_dest,
            'next_hop_type': next_hop_type
        })
    
    return route_table_dict

def find_destination_vcn(cidr_block, df_vcns):
    try:
        return df_vcns.loc[df_vcns['CIDR_Block'] == cidr_block, 'VCN_Name'].values[0]
    except IndexError:
        return "Unknown_VCN"


route_tables = extract_route_tables(excel_file_path)

#--------------------------------------------------------------------------------------------------------------------------

def rt_associations(wb):
    rt_sheet = wb["RT_Attachment"]
    associations = []

    for row in rt_sheet.iter_rows(min_row=2, values_only=True):
        rt_name, subnet_name = row
        associations.append({"route_table": rt_name, "subnet": subnet_name})

    return associations

rt_assocs = rt_associations(wb)

#--------------------------------------------------------------------------------------------------------------------------

template_data = {
    "resource_group": resource_group,
    "region": region,
    "hub_vcn_name": hub_vcn_name,
    "spoke_vcns": spoke_vcns,
    "route_tables": route_tables,
    "route_table_associations": rt_assocs

}

env = Environment(
    loader=FileSystemLoader(os.path.dirname(template_file_path)),
    autoescape=select_autoescape(['j2'])
)

template = env.get_template(os.path.basename(template_file_path))

rendered_content = template.render(template_data)

with open(output_file_path, 'w') as f:
    f.write(rendered_content)

print("Fichier variables.tfvars généré avec succès.")
